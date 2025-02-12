from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook

# Excel file setup
excel_file = "upload_results.xlsx"
sheet_name = "Results"

# Check if Excel file exists, if not, create it
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(["Test Case", "Upload Message", "Status"])  # Headers
    wb.save(excel_file)

# Initialize WebDriver
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)

# Open the image upload website
driver.get("https://tus.io/demo.html")
print("Website launched successfully")

# Ensure the file exists before uploading
image_path = os.path.abspath("testingfile.jpg")
if os.path.exists(image_path):
    print("✅ File found:", image_path)
else:
    print("❌ File NOT found:", image_path)
    driver.quit()
    exit()

# Maximize window
driver.maximize_window()
print("Website maximized")

# Get the file input element
file_input = driver.find_element(By.XPATH, "//input[@type='file']")
print("Found the input element")

# Scroll to make the element visible
driver.execute_script("window.scrollBy(0, 500);")
time.sleep(1)

# Highlight the element
driver.execute_script("arguments[0].style.border='3px solid red'", file_input)
print("Highlighted the upload element")

# Upload the image
file_input.send_keys(image_path)
print("Image uploaded successfully")

# ✅ **Wait for the success message to appear**
try:
    success_message_element = WebDriverWait(driver, 30).until(
        EC.visibility_of_element_located((By.XPATH, "//p[text()='The upload is complete!']"))
    )
    print("✅ Success message found!")

    success_message = success_message_element.text
    print("✅ Upload Message:", success_message)

    # ✅ **Verify and update in Excel**
    expected_message = "The upload is complete!"
    status = "Pass" if success_message == expected_message else "Fail"

    # Open existing Excel file and write results
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]
    ws.append(["Image Upload Test", success_message, status])
    wb.save(excel_file)

    print(f"✅ Result saved in Excel: {status}")

except Exception as e:
    print("❌ No success message found:", e)

    # Log failure in Excel
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]
    ws.append(["Image Upload Test", "No message found", "Fail"])
    wb.save(excel_file)

# Close the browser
driver.quit()
